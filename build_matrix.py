"""
Generate CSV and XLSX deliverables from the canonical KPI definitions.

Outputs:
  - data/kpi_comparison_matrix.csv       (the citable dataset)
  - data/kpi_comparison_matrix.xlsx      (formatted version for human reading)
"""

import csv
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from kpi_definitions import KPIS

OUT_DIR = Path("data")
OUT_DIR.mkdir(exist_ok=True)

# CSV column order matters: most-stable identifiers first, narrative last
CSV_COLUMNS = [
    "id", "name", "tier", "family",
    "formula", "units", "domain",
    "source", "assumptions",
    "limitations_in_BWk_BSk", "regional_extension_rationale",
    "references",
]

# ─────────────────────────────────────────────────────────────────────────────
# CSV — the citable dataset
# ─────────────────────────────────────────────────────────────────────────────
csv_path = OUT_DIR / "kpi_comparison_matrix.csv"
with csv_path.open("w", newline="", encoding="utf-8") as f:
    writer = csv.DictWriter(f, fieldnames=CSV_COLUMNS, quoting=csv.QUOTE_ALL)
    writer.writeheader()
    for kpi in KPIS:
        writer.writerow({col: kpi[col] for col in CSV_COLUMNS})
print(f"Wrote {csv_path}  ({len(KPIS)} rows)")

# ─────────────────────────────────────────────────────────────────────────────
# XLSX — formatted version
# ─────────────────────────────────────────────────────────────────────────────
TIER_COLORS = {
    "CORE":        "C6EFCE",   # green
    "EXTENSION":   "FFEB9C",   # yellow
    "EXPLORATORY": "F2DCDB",   # pink
}
HEADER_FILL = PatternFill("solid", start_color="1F3864")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
BODY_FONT   = Font(name="Calibri", size=10)
BOLD_FONT   = Font(name="Calibri", size=10, bold=True)
ALT_FILL    = PatternFill("solid", start_color="F5F5F5")
THIN        = Side(border_style="thin", color="BFBFBF")
BORDER      = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

WRAP_CENTER = Alignment(wrap_text=True, vertical="center", horizontal="left")
HEADER_ALN  = Alignment(wrap_text=True, vertical="center", horizontal="center")
ID_ALN      = Alignment(wrap_text=False, vertical="center", horizontal="center")

wb = Workbook()

# ── Sheet 1: README ──────────────────────────────────────────────────────────
ws_readme = wb.active
ws_readme.title = "README"

readme_rows = [
    ("KPI Comparison Matrix v1.0", "header"),
    ("Arid-climate resilience indicators for residential buildings — South American context", "subhead"),
    ("", "blank"),
    ("Author", "Dr. Arq. Gustavo Javier Barea Paci  ·  INAHE-CONICET Mendoza  ·  gbarea@mendoza-conicet.gob.ar"),
    ("Repository", "https://github.com/<user>/kpi-comparison-matrix  (placeholder — to be created)"),
    ("DOI", "10.5281/zenodo.XXXXXXX  (placeholder — to be assigned on Zenodo release)"),
    ("Licence", "Creative Commons Attribution 4.0 International (CC BY 4.0)"),
    ("Version", "1.0  ·  released YYYY-MM-DD"),
    ("Companion library", "arid-resilience-kpis  (Python package implementing the 5 core KPIs)"),
    ("", "blank"),
    ("Tier system", "header"),
    ("CORE", "5 indicators computed on 100 % of cases of the 2026–2029 plan."),
    ("EXTENSION", "7 indicators computed when data quality permits; reinforce the core findings."),
    ("EXPLORATORY", "5 indicators computed opportunistically; potential contributions subject to availability."),
    ("", "blank"),
    ("Contents of the matrix sheet", "header"),
    ("id", "Short stable identifier."),
    ("name", "Full descriptive name."),
    ("tier", "CORE / EXTENSION / EXPLORATORY."),
    ("family", "Conceptual family the indicator belongs to."),
    ("formula", "Formal definition in plain text (Unicode subscripts kept ASCII for portability)."),
    ("units", "Physical unit; 'dimensionless' when applicable."),
    ("domain", "What the indicator measures and why it matters."),
    ("source", "Primary literature or this work."),
    ("assumptions", "Assumptions required for the indicator to be valid."),
    ("limitations_in_BWk_BSk", "Documented limitations in arid cold/semi-arid climates of the Köppen-Geiger classification."),
    ("regional_extension_rationale", "Justification for using or extending the indicator in our regional context."),
    ("references", "Bibliographic references (semicolon-separated)."),
    ("", "blank"),
    ("How to cite", "header"),
    ("Citation", "Barea Paci, G. J. (2026). KPI Comparison Matrix v1.0 — arid-climate resilience indicators for residential buildings. Zenodo. https://doi.org/10.5281/zenodo.XXXXXXX"),
    ("", "blank"),
    ("Acknowledgement", "header"),
    ("Plan of work 2026–2029", "This dataset is the first deliverable of the IEA EBC Annex 80-aligned framework being developed at INAHE-CONICET for arid South American climates (BWk/BSk). It feeds the matrix of objectives OE1 and OE2 of the 2026–2029 plan and is consumed by the companion Python library arid-resilience-kpis."),
]

# layout
ws_readme.column_dimensions["A"].width = 22
ws_readme.column_dimensions["B"].width = 110
row = 1
for label, value in readme_rows:
    if value == "header":
        ws_readme.cell(row=row, column=1, value=label).font = Font(name="Calibri", size=14, bold=True, color="1F3864")
        row += 1
    elif value == "subhead":
        ws_readme.cell(row=row, column=1, value=label).font = Font(name="Calibri", size=11, italic=True, color="595959")
        row += 1
    elif value == "blank":
        row += 1
    else:
        c1 = ws_readme.cell(row=row, column=1, value=label)
        c2 = ws_readme.cell(row=row, column=2, value=value)
        c1.font = BOLD_FONT
        c2.font = BODY_FONT
        c2.alignment = Alignment(wrap_text=True, vertical="top")
        ws_readme.row_dimensions[row].height = max(20, 16 * (1 + len(str(value)) // 100))
        row += 1

# ── Sheet 2: matrix ──────────────────────────────────────────────────────────
ws = wb.create_sheet("KPI matrix")

# Display columns and widths (in xlsx character units)
DISPLAY_COLUMNS = [
    ("id",                              8),
    ("name",                            32),
    ("tier",                            14),
    ("family",                          24),
    ("formula",                         48),
    ("units",                           14),
    ("domain",                          40),
    ("source",                          28),
    ("assumptions",                     38),
    ("limitations_in_BWk_BSk",          42),
    ("regional_extension_rationale",    42),
    ("references",                      40),
]

# Headers
for j, (col_id, width) in enumerate(DISPLAY_COLUMNS, start=1):
    cell = ws.cell(row=1, column=j, value=col_id)
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = HEADER_ALN
    cell.border = BORDER
    ws.column_dimensions[get_column_letter(j)].width = width
ws.row_dimensions[1].height = 28

# Sort rows: CORE first, then EXTENSION, then EXPLORATORY (within each, by id)
TIER_ORDER = {"CORE": 0, "EXTENSION": 1, "EXPLORATORY": 2}
sorted_kpis = sorted(KPIS, key=lambda k: (TIER_ORDER[k["tier"]], k["id"]))

for i, kpi in enumerate(sorted_kpis, start=2):
    tier_color = TIER_COLORS[kpi["tier"]]
    for j, (col_id, _) in enumerate(DISPLAY_COLUMNS, start=1):
        value = kpi[col_id]
        cell = ws.cell(row=i, column=j, value=value)
        cell.font = BOLD_FONT if col_id in ("id", "name") else BODY_FONT
        if col_id in ("id", "tier"):
            cell.fill = PatternFill("solid", start_color=tier_color)
            cell.alignment = ID_ALN
        else:
            cell.alignment = WRAP_CENTER
            if i % 2 == 1:
                cell.fill = ALT_FILL
        cell.border = BORDER
    ws.row_dimensions[i].height = 95

# Freeze header + id column
ws.freeze_panes = "C2"

# Auto filter on header row
ws.auto_filter.ref = f"A1:{get_column_letter(len(DISPLAY_COLUMNS))}{len(sorted_kpis)+1}"

# ── Sheet 3: Tier legend ─────────────────────────────────────────────────────
ws_legend = wb.create_sheet("Tier legend")
ws_legend.column_dimensions["A"].width = 18
ws_legend.column_dimensions["B"].width = 14
ws_legend.column_dimensions["C"].width = 90

ws_legend.cell(row=1, column=1, value="Tier").font = HEADER_FONT
ws_legend.cell(row=1, column=2, value="Colour").font = HEADER_FONT
ws_legend.cell(row=1, column=3, value="Operational meaning in the 2026–2029 plan").font = HEADER_FONT
for j in range(1, 4):
    ws_legend.cell(row=1, column=j).fill = HEADER_FILL
    ws_legend.cell(row=1, column=j).alignment = HEADER_ALN
    ws_legend.cell(row=1, column=j).border = BORDER

legend_rows = [
    ("CORE",        "green",  "Computed on 100 % of cases. Forms the committed nucleus of the plan; reported for every house, every typology, every scenario."),
    ("EXTENSION",   "yellow", "Computed where the data quality permits (longer monitoring records, calibrated models). Reinforces core findings; reported for the subset of houses that meet the data-quality threshold."),
    ("EXPLORATORY", "pink",   "Computed opportunistically. Constitutes potential additional contributions; reported when a paper or project explicitly requires the indicator and the necessary data is available."),
]
for i, (tier, colour, meaning) in enumerate(legend_rows, start=2):
    ws_legend.cell(row=i, column=1, value=tier).font = BOLD_FONT
    ws_legend.cell(row=i, column=1).fill = PatternFill("solid", start_color=TIER_COLORS[tier])
    ws_legend.cell(row=i, column=1).alignment = ID_ALN
    ws_legend.cell(row=i, column=2, value=colour).alignment = ID_ALN
    ws_legend.cell(row=i, column=2).fill = PatternFill("solid", start_color=TIER_COLORS[tier])
    ws_legend.cell(row=i, column=3, value=meaning).alignment = WRAP_CENTER
    ws_legend.cell(row=i, column=3).font = BODY_FONT
    for j in range(1, 4):
        ws_legend.cell(row=i, column=j).border = BORDER
    ws_legend.row_dimensions[i].height = 60

# ── Save ─────────────────────────────────────────────────────────────────────
xlsx_path = OUT_DIR / "kpi_comparison_matrix.xlsx"
wb.save(xlsx_path)
print(f"Wrote {xlsx_path}")

# Stats
print("\nSummary:")
for tier in ("CORE", "EXTENSION", "EXPLORATORY"):
    n = sum(1 for k in KPIS if k["tier"] == tier)
    print(f"  {tier:12s} : {n} indicators")
